from random import sample
import requests
from bs4 import BeautifulSoup
from lxml import etree
from openpyxl import Workbook


def get_courses_list(num_of_links):
    links = []
    courses_url = 'https://www.coursera.org/sitemap~www~courses.xml'
    request = requests.get(courses_url)
    root = etree.fromstring(request.content)
    for item in root.iter('{*}loc'):
        links.append(item.text)
    random_links = sample(links, num_of_links)
    return random_links


def get_course_details(link):
    course_response = requests.get(link)
    return course_response


def get_course_info(random_links):
    course_info = []
    for link in random_links:
        course_response = get_course_details(link)
        soup = BeautifulSoup(course_response.text, 'lxml')
# язык, ближайшую дату начала, количество недель и среднюю оценку
        course_title = soup.find('div', {'class': 'title'})
        course_lang = soup.find('div', {'class': 'language-info'})
        course_rating = soup.find(
            'div',
            {'class': 'ratings-text bt3-visible-xs'}
        )
        course_start = soup.find(
            'div',
            {'class': 'startdate rc-StartDateString caption-text'}
            )
        if course_rating is not None:
            course_rating = course_rating.text
        course_duration = soup.findAll('div', {'class': 'week-heading'})
        if course_duration is not None:
            course_duration = len(course_duration)
        course_info.append({
                            'course title': course_title.text,
                            'course lang': course_lang.text,
                            'course rating': course_rating,
                            'course duration': course_duration,
                            'course start': course_start.text
        }
        )
    return(course_info)


def output_courses_info_to_xlsx(filepath, course_info):
    wb = Workbook()
    sheet = wb.active
    sheet_head = (
        'Название курса',
        'Языки',
        'Дата начала',
        'Длительность курса',
        'Рейтинг'
        )
    for head_num, head in enumerate(sheet_head, 1):
        sheet.cell(row=1, column=head_num).value = head
    for course_num, course in enumerate(course_info, 3):
        sheet.cell(row=course_num, column=1).value = course['course title']
        sheet.cell(row=course_num, column=2).value = course['course lang']
        sheet.cell(row=course_num, column=3).value = course['course start']
        sheet.cell(row=course_num, column=4).value = course['course duration']
        sheet.cell(row=course_num, column=5).value = course['course rating']
    wb.save(filepath)


if __name__ == '__main__':
    num_of_links = 20
    filepath = 'course_info.xlsx'
    random_links = get_courses_list(num_of_links)
    course_info = get_course_info(random_links)
    output_courses_info_to_xlsx(filepath, course_info)
